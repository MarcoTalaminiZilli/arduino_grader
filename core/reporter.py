import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

class ExcelReporter:
    @staticmethod
    def generate_report(results: list, output_path: str):
        """
        Gera um relatório Excel a partir dos resultados da avaliação.
        
        Cada item da lista `results` deve conter os campos:
        - aluno_id (str)
        - exercicio (str)
        - has_essential (bool)
        - final_score (float)
        - feedback (str)
        """
        # Monta a estrutura de dados incluindo o Feedback da IA
        data = []
        for item in results:
            data.append({
                "ID Aluno": item.get("aluno_id", "N/A"),
                "Exercício": item.get("exercicio", "N/A"),
                "Estruturas Básicas": "OK" if item.get("has_essential") else "Ausente",
                "Nota Final": item.get("final_score", 0.0),
                "Feedback da IA": item.get("feedback", "Sem observações.")
            })

        df = pd.DataFrame(data)

        # Exporta para o Excel aplicando formatação visual com openpyxl
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Resultados")
            workbook = writer.book
            worksheet = writer.sheets["Resultados"]

            # Estilo do Cabeçalho (Azul escuro com texto branco e negrito)
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Formatação e largura das colunas
            for col in worksheet.columns:
                col_name = col[0].value
                col_letter = get_column_letter(col[0].column)

                if col_name == "Feedback da IA":
                    # Define largura fixa e ativa quebra automática para frases longas
                    worksheet.column_dimensions[col_letter].width = 50
                    for cell in col[1:]:
                        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
                else:
                    # Ajusta largura automaticamente baseada no maior texto da coluna
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
                    for cell in col[1:]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

            # Aumenta a altura das linhas de dados para acomodar os feedbacks multilinha
            for row in range(2, len(df) + 2):
                worksheet.row_dimensions[row].height = 45